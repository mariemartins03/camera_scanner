"""
Exportação de resultados para Excel (.xlsx).
"""
from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from config import Config
from fabricantes import CameraInfo

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
_HEADER_BG = "1E3A5F"       # azul escuro
_HEADER_FG = "FFFFFF"       # branco
_ALT_ROW_BG = "EEF3FA"      # azul muito claro para linhas alternadas
_OK_FG = "1A7F3C"           # verde para status OK
_ERR_FG = "C0392B"          # vermelho para status Erro
_BORDER_COLOR = "CCCCCC"

def _thin_border():
    side = Side(style="thin", color=_BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)

# ---------------------------------------------------------------------------
# Exportação
# ---------------------------------------------------------------------------

COLUNAS = [
    ("IP",          "ip",          16),
    ("Fabricante",  "fabricante",  16),
    ("Modelo",      "modelo",      24),
    ("MAC",         "mac",         20),
    ("Status",      "status",      12),
    ("Observação",  "erro",        36),
]


def exportar_excel(cameras: Sequence[CameraInfo], nome_arquivo: str | None = None) -> str:
    """
    Gera um arquivo Excel com os resultados da varredura.

    Args:
        cameras: Lista de CameraInfo a exportar.
        nome_arquivo: Nome do arquivo (sem caminho). Se None, gera automaticamente.

    Returns:
        Caminho absoluto do arquivo gerado.
    """
    if nome_arquivo is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"cameras_{ts}.xlsx"

    caminho = os.path.join(Config.EXPORTS_DIR, nome_arquivo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Câmeras"

    # Linha de título
    ws.merge_cells("A1:F1")
    titulo_cell = ws["A1"]
    titulo_cell.value = f"Camera Scanner — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    titulo_cell.font = Font(bold=True, size=13, color=_HEADER_FG)
    titulo_cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Cabeçalhos
    header_row = 2
    for col_idx, (label, _, largura) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True, color=_HEADER_FG)
        cell.fill = PatternFill("solid", fgColor="2E5F9E")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.row_dimensions[header_row].height = 20

    # Dados
    for row_idx, cam in enumerate(cameras, start=3):
        is_alt = (row_idx % 2 == 0)
        row_data = [getattr(cam, campo) for _, campo, _ in COLUNAS]

        for col_idx, valor in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor or "")
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = _thin_border()

            if is_alt:
                cell.fill = PatternFill("solid", fgColor=_ALT_ROW_BG)

        # Colorir coluna Status
        status_cell = ws.cell(row=row_idx, column=5)
        if cam.status == "OK":
            status_cell.font = Font(bold=True, color=_OK_FG)
        elif cam.status == "Erro":
            status_cell.font = Font(bold=True, color=_ERR_FG)

    # Rodapé com total
    ultimo_dado = len(cameras) + 2
    rodape_row = ultimo_dado + 2
    ws.merge_cells(f"A{rodape_row}:F{rodape_row}")
    rodape_cell = ws.cell(row=rodape_row, column=1)
    rodape_cell.value = f"Total: {len(cameras)} câmera(s) encontrada(s)"
    rodape_cell.font = Font(italic=True, color="666666")
    rodape_cell.alignment = Alignment(horizontal="right")

    # Congelar cabeçalho
    ws.freeze_panes = "A3"

    wb.save(caminho)
    logger.info("Excel exportado: %s (%d câmeras)", caminho, len(cameras))
    return caminho
